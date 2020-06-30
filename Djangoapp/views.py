from django.contrib.admin.helpers import AdminForm
from django.db.models import Avg
from django.shortcuts import render, redirect
from tem.forms import EmployeeForm, DmForm, LoginForm, COForm, MappingForm, PosForm,\
    AdminLoginForm, TsForm, EsForm, WeightForm, TeacherForm, TeachesForm
from tem.models import Employee, PO, CO, Upload_Int, Upload_Ext, ExamScheme, Assesment, Admin, Weights, CoPoMapp, \
    Report, DeliveryMethods, TeachingScheme
from django.http import HttpResponse
from django.views.generic import View
from .utils import render_to_pdf
from django.views.decorators.cache import cache_control
import openpyxl

cnt=1
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def blog(request):
    return render(request, "blog.html")


def courses1(request):
    return render(request, "courses1.html")


def dbms(request):
    return render(request, "dbms.html")


def home(request):
    return render(request, "home.html")


def programs(request):
    return render(request, "programs.html")


def os(request):
    return render(request, "os.html")


def hci(request):
    return render(request, "hci.html")


def toc(request):
    return render(request, "toc.html")


def sepm(request):
    return render(request, "sepm.html")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def course(request):
    s1 = request.session['course']
    global v1, cnt, s2, flag, v2
    s2 = s1
    if request.method == "POST":
        form = COForm(request.POST)
        if form.is_valid:
            try:
                temp = form.save(commit=False)
                s2 = s2 + '.' + str(cnt)  # 1441_2015.1 ,.2, .3 .....
                temp.CO_id = s2  # assign that value
                temp.save()
                cnt = cnt + 1
                
                v1 = s1
                v2 = s2[:-2] + '.' + str(cnt-1)
                print(v2, cnt)
                fo = CO.objects.all()
                return render(request, "co.html", {'fo': fo , 'v1': v1, 'v2': v2,'cnt': cnt})
            except:
                print('unsaved')
                pass
        else:
            print('not validated')
    else:
        form = COForm()
        v1 = s2
        v2 = s2 + '.' + str(cnt)
    return render(request, "courses.html", {'form': form, 'v1': v1, 'v2': v2})


def skipfn(request):
    skipflag = 1
    return redirect('/mapping')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def mapping(request):
    global v1, v2, s2, cnt, flag
    s1 = request.session['course']
    s2 = s1
    if request.method == "POST":
        form = MappingForm(request.POST)
        if form.is_valid:
            try:
                temp = form.save(commit=False)
                s2 = s2 + '.' + str(cnt)
                cos = CO.objects.all()
                flag = 0
                for i in cos:
                    if i.CO_id == s2:
                        obj = CO.objects.get(CO_id=s2)
                        temp.CO_id = obj
                        temp.save()
                        flag = 1
                        break
                if flag == 1:
                    cnt = cnt + 1
                    form = MappingForm()
                    v1 = s1
                    s2 = s2[:-2] + '.' + str(cnt)
                    v3 = obj.Outcome
                    print(v3,'1')
                    return render(request, "blog.html", {'form': form, 'v1': s1, 'v2': s2,'v3': v3})
                else:
                    print('now done with mapping')
                    return redirect('/pricing')
            except:
                print('unsaved')
                pass
        else:
            print('not validated')
    else:
        cnt = 1
        form = MappingForm()
        v1 = s1
        s2 = s1 + '.' + '1'
        obj  = CO.objects.get(CO_id=s2)
        v3 = obj.Outcome
        print(s2)
    return render(request, "blog.html", {'form': form, 'v1': s1, 'v2': s2,'v3': v3})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def contact(request):
    v = request.session['course']
    return render(request, "contact.html", {'v': v})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def pricing():
    return redirect('/show')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def index1(request):
    return render(request,"index1.html")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def emp(request):
    if request.method == "POST":
        request.session['course'] = request.POST.get('course')
        request.session['password'] = request.POST.get('password')
        if request.POST.get('password') == request.POST.get('re-password'):
            form = EmployeeForm(request.POST)
            if form.is_valid():
                try:
                    form.save()
                    return redirect('/access')
                except:
                    pass
            else:
                print(1)
                return redirect('/')
        else:
            print(2)
            return redirect('/')
    else:
        form = EmployeeForm()
        return render(request, "home.html", {'form': form})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def loginpage(request):
    data2 = Employee.objects.all()
    if request.method == "POST":
        request.session['course'] = request.POST.get('course')
        request.session['password'] = request.POST.get('password')
        if request.POST.get('password') == request.POST.get('re-password'):
            form1 = LoginForm(request.POST)
            if form1.is_valid():
                print(3)
                ext = form1.cleaned_data
                for i in data2:
                    codev = ext.get("course")
                    paswd = ext.get("password")
                    if i.course == codev and i.password == paswd:
                        return redirect('/access')
        
                return redirect('/')
        else:
            return redirect('/')
    else:
        form1 = EmployeeForm()
        return render(request, "index.html", {'form1': form1})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def access_session(request):
    data1 = {
        'x': request.session['course'],
        'y': request.session['password']
    }
    form = DmForm()
    return render(request, "index1.html", {'data1': data1,'form': form })


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dm(request):  # -------------logic of dmform with foreign keys
    global v
    if request.method == "POST":
        form = DmForm(request.POST)
        if form.is_valid():
            try:
                s1 = request.session['course']# 1441_2015 or any code like that coming from login
                obj2 = Employee.objects.get(course=s1)  # object where Code_patt == logined code
                obj = form.save(commit=False)  # creating temporary object
                obj.Code_patt = obj2  # assign that value
                obj.save()  # save it
                print('saved')
                return redirect('/course')
            except:
                print('unsaved')
                return redirect('/dm')
                pass
        else:
            print('not validated')
    else:
        return redirect('/access')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def show(request):
    pos = {
        'pos1': PO.objects.get(id="1"),
        'pos2': PO.objects.get(id="2"),
        'pos3': PO.objects.get(id="3"),
        'pos4': PO.objects.get(id="4"),
        'pos5': PO.objects.get(id="5"),
        'pos6': PO.objects.get(id="6"),
        'pos7': PO.objects.get(id="7"),
        'pos8': PO.objects.get(id="8"),
        'pos9': PO.objects.get(id="9"),
        'pos10': PO.objects.get(id="10"),
        'pos11': PO.objects.get(id="11"),
        'pos12': PO.objects.get(id="12")
    }
    s1 = request.session['course']
    v = s1
    return render(request,"pricing.html",{'pos': pos,'v':v})
    
    
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def delete_session(request):
    try:
        del request.session['course']
        del request.session['password']
    except KeyError:
        pass
    return redirect('/')


class GeneratePdf(View):
    print('hjjhj')
    def get(self, request, *args, **kwargs):
        print(2)
        s1 = request.session['course']
        coursedetails = Employee.objects.get(course=s1)
        DM = DeliveryMethods.objects.filter(Code_patt=s1)
        cos = CO.objects.filter(CO_id__icontains=s1)
        pos = PO.objects.all()
        #map = CoPoMapp.objects.filter(CO_id_id__icontains=s1)
        data = {
            'coursedetails': coursedetails,
            'DM' : DM,
            'CO' : cos,
            'po' : pos,
        }
        print(3)
        pdf = render_to_pdf('pdf/invoice.html', data)
        print(4)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Invoice_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            download = request.GET.get("download")
            if download:
                content = "attachment; filename='%s'" % (filename)
            response['Content-Disposition'] = content
            return response
        return HttpResponse("Not found")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def ind(request):
    if "POST" == request.method:
        excel_file = request.FILES["excel_file"]
        
        wb = openpyxl.load_workbook(excel_file)
        
        # getting all sheets
        sheets = wb.sheetnames
        # print(sheets)
        
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # print(worksheet)
        
        # getting active sheet
        active_sheet = wb.active
        # print(active_sheet)
        
        # reading a cell
        # print(worksheet["A1"].value)
        
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        arr = [0,0,0,0,0,0,0,0,0,0,0]
        for row in worksheet.iter_rows():
            row_data = list()
            cnt = 0
            for cell in row:
                row_data.append(str(cell.value))
                arr[cnt] = cell.value
                cnt = cnt + 1
            x= request.session['course']
            obj1 = Employee.objects.get(course=x)
            p = Upload_Int(Code_patt=obj1,A1=arr[0], A2=arr[1], A3=arr[2], A4=arr[3], A5=arr[4],ut1=arr[5], ut2=arr[6], ut3=arr[7], ut4=arr[8], ut5=arr[9],tw=arr[10])
            p.save()
            excel_data.append(row_data)
        
        # form.save()
        return render(request, 'contact.html')
    else:
        return render(request, 'contact.html', {})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def indi(request):
    if "POST" == request.method:
        excel_file = request.FILES["excel_file"]
        
        wb = openpyxl.load_workbook(excel_file)
        
        # getting all sheets
        sheets = wb.sheetnames
        # print(sheets)
        
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # print(worksheet)
        
        # getting active sheet
        active_sheet = wb.active
        # print(active_sheet)
        
        # reading a cell
        # print(worksheet["A1"].value)
        
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        arr = [0,0,0,0,0]
        for row in worksheet.iter_rows():
            row_data = list()
            cnt = 0
            for cell in row:
                row_data.append(str(cell.value))
                arr[cnt] = cell.value
                cnt = cnt + 1
            
            x1 = request.session['course']
            obj2= Employee.objects.get(course=x1)
            p = Upload_Ext(Code_patt=obj2, insem=arr[0], practicals=arr[1], TW =arr[2],endsem=arr[3])
            p.save()
            excel_data.append(row_data)
        
        # form.save()
        return render(request, 'contact.html')
    else:
        return render(request, 'contact.html', {})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def result(request):
    print('calculating..')
    global no_of_students, y, p
    no_of_students = 0
    code = request.session['course']
    temp = Upload_Int.objects.all()
  
    for i in temp:
        print(i.Code_patt, code)
        if i.Code_patt == code:
            no_of_students = no_of_students + 1

    print(no_of_students)
    cos = CO.objects.all()
    for i in cos:
        print(i.CO_id)
        if code == i.CO_id[:-2]:  # logged in course code and its corrosponding CO
            assignments = i.assignments  # 1,2,3
            unittests = i.units  # 1,2,2
            avgint = 0
            avgext = 0
            avg = [0, 0, 0, 0, 0, 0]
            avgut = [0, 0, 0, 0, 0, 0]
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('A1'))  # avg of A1 of 1441_2015
            avg[1] = x['A1__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('A2'))
            avg[2] = x['A2__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('A3'))
            avg[3] = x['A3__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('A4'))
            avg[4] = x['A4__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('A5'))
            avg[5] = x['A5__avg']
            print('121212')
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('ut1'))  # avg of unit test 1 of 1441_2015
            avgut[1] = x['ut1__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('ut2'))
            avgut[2] = x['ut2__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('ut3'))
            avgut[3] = x['ut3__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('ut4'))
            avgut[4] = x['ut4__avg']
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('ut5'))
            avgut[5] = x['ut5__avg']
            print('22')
            x = Upload_Int.objects.filter(Code_patt=code).aggregate(Avg('tw'))
            avgtw = x['tw__avg']  # out of exam scheme
            x = Upload_Ext.objects.filter(Code_patt=code).aggregate(Avg('practicals'))
            avgpract = x['practicals__avg']
            x = Upload_Ext.objects.filter(Code_patt=code).aggregate(Avg('insem'))
            avginsem = x['insem__avg']
            x = Upload_Ext.objects.filter(Code_patt=code).aggregate(Avg('endsem'))
            avgendsem = x['endsem__avg']
            print('33')
            obj = ExamScheme.objects.all()  # getting details of exam scheme i.e. uot of marks
            for l in obj:
                print(l.Code_patt_id, code)
                if l.Code_patt_id == code:
                    y = l.termwork
                    avgtw = (100 * avgtw) / y
                    print(avgtw)
                    y = l.Practicle
                    avgpract = (100 * avgpract) / y
                    y = l.onlineinsem
                    avginsem = (100 * avginsem) / y
                    y = l.endsem
                    avgendsem = (100 * avgendsem) / y
                    break
            print(4)

            j = 0
            summ = 0
            cnt = 0

            while j < len(assignments):
                if assignments[j] != ',':
                    cnt = cnt + 1
                    t = int(assignments[j])
                    summ = summ + avg[t] * 10  # avg percwnt of one assignment

                j = j + 1

            avgassignments = summ / cnt
            summ = 0
            j = 0
            cnt = 0
            f1 = f2 = f3 = 0
            while j < len(unittests):
                if unittests[j] != ',':
                    cnt = cnt + 1
                    t = int(unittests[j])
                    summ = summ + avgut[t] * 10  # avg percwnt of one unit test

                j = j + 1
            print(6)
            avgunits = summ / cnt
            arr = [0, 0, 0, 0, 0, 0, 0, 0]
            z = 0
            if not i.insem:
                avginsem = 0
            if not i.endsem:
                avgendsem = 0
            if not i.practicle:
                avgpract = 0
            if not i.TW:
                avgtw = 0
                avgint = (avgassignments + avgunits) / 2
            else:
                avgint = (avgassignments + avgunits + avgtw) / 3
            print(7)
            arr[0] = avginsem
            arr[1] = avgendsem
            arr[2] = avgpract
            o = 0
            div = 0
            while o < 3:
                if arr[o] != 0:
                    div = div + 1
                o = o + 1
            print(8)
            avgext = (avginsem + avgendsem + avgpract) / div
            getwt = Weights.objects.filter(Code_patt=code)
            for o in getwt:
                avgext = avgext * o.ext_weight
                avgint = avgint * o.int_weight
            print(9)
            final = avgext + avgint
            if final >= 80:
                lvl = 3
            elif final >= 70:
                lvl = 2
            elif final >= 60:
                lvl = 1
            else:
                lvl = 0
            obj1 = CO.objects.get(CO_id=i.CO_id)
            print('saving')
            p = Assesment(CO_id=obj1, assignment=avgassignments, insem=avginsem, endsem=avgendsem, practicle=avgpract,
                          termwork=avgtw, unittests=avgunits, assessment=final, level=lvl)
            p.save()  # assessment of that CO

    return redirect('/display2')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def display2(request):
    form = Assesment.objects.all()
    s1 = request.session['course']
    return render(request, "report.html", {'form': form,'s1':s1})


def courseinfo(request):
    if request.method == "POST":
        form2 = EmployeeForm(request.POST)
        if form2.is_valid():
            try:
                form2.save()
                return redirect('/adminhomepage')
            except:
                pass
    else:
        form2 = EmployeeForm()
    return render(request, "entry.html", {'form2': form2})


def adminhomepage(request):
    return render(request, "adminhome.html")


def adminentry(request):
    if request.method == "POST":
        form = AdminForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('/adminlogin')
            except:
                pass
    else:
        form = AdminForm()
    return render(request, "adminentry.html", {'form': form})


def adminlogin(request):
    if request.method == "POST":
        form = AdminLoginForm(request.POST)
        data2 = Admin.objects.all()
        if form.is_valid():
            ext = form.cleaned_data
            for i in data2:
                uname = ext.get("username")
                paswd = ext.get("password")
                if i.username == uname and i.password == paswd:
                    return redirect('/adminhomepage')
                else:
                    return redirect('/adminentry')

            return redirect('/')
        else:
            form = AdminLoginForm()
            return render(request, "adinlogin.html", {'form': form})


def poentry(request):
    if request.method == "POST":
        form = PosForm(request.POST)
        if form.is_valid:
            try:
                form.save()
                form = PosForm()
                return render(request, "PO.html", {'form': form})
            except:
                print('unsaved')
                pass
        else:
            print('not validated')
    else:
        form = PosForm()
    return render(request, "PO.html", {'form': form})


def coursescheme(request):
    global s1
    s1 = Employee.ename
    if request.method == "POST":
        form = TsForm(request.POST)
        form2 = EsForm(request.POST)
        if form.is_valid() and form2.is_valid():
            try:
                f1 = form.save(commit=False)
                f2 = form2.save(commit=False)
                s1 = Employee.ename  # 1441_2015 or any code like that coming from login
                obj2 = Employee.objects.get(Code_patt=s1)
                f1.Code_patt = obj2
                f2.Code_patt = obj2
                f1.save()
                f2.save()
                print('hhhh')
                return redirect('/homepage2')
            except:
                print('unsaved')
                pass
        else:
            print('not validated')
    else:
        form = TsForm()
        form2 = EsForm()
    print(s1)
    return render(request, "schemes.html", {'form': form, 'form2': form2, 's1': s1})


def weights(request):
    if request.method == "POST":
        form = WeightForm(request.POST)
        if form.is_valid:
            try:
                s1 = request.session['course']  # 1441_2015 or any code like that coming from login
                obj2 = Employee.objects.get(course=s1)
                print('1')
                obj = form.save(commit=False)
                x = obj.ext_weight
                print(x)
                print(1.00000000000 - x)
                y = 1 - x
                y = y * 10
                y = round(y) / 10
                obj.int_weight = y
                obj.Code_patt = obj2
                obj.save()
                print(obj)
                return redirect('/result')
            except:
                print('unsaved')
                pass
        else:
            print('not validated')
    else:
        print('jkhjuhhhhhhhhhhhhgb')
        form = WeightForm()
        s1 = request.session['course']
    return render(request, "attainment.html", {'form': form, 's1': s1})


def teacherentry(request):
    if request.method == "POST":
        form2 = TeacherForm(request.POST)
        if form2.is_valid():
            try:
                form2.save()
                return redirect('/adminhomepage')
            except:
                print('unsaved')
                pass
    else:
        print('not validated')
        form2 = TeacherForm()
    return render(request, "teacherentry.html", {'form2': form2})


def teacher_alloc(request):
    if request.method == "POST":
        form2 = TeachesForm(request.POST)
        if form2.is_valid():
            try:
                form2.save()
                return redirect('/adminhomepage')
            except:
                pass
    else:
        form2 = TeachesForm()
    return render(request, "teacher_alloc.html", {'form2': form2})


def report_gen(request):
    # Co po mapping , assessment
    code = request.session['course']
    results = Assesment.objects.all()
    mappings = CoPoMapp.objects.all()
    for i in results:
        if i.CO_id_id[:-2] == code:   # got assessment of a CO of logged in course
            for j in mappings:
                if j.CO_id_id == i.CO_id_id:
                    r1 = (i.assignment * j.VPO1) / 3
                    r2 = (i.assignment * j.VPO2) / 3
                    r3 = (i.assignment * j.VPO3) / 3
                    r4 = (i.assignment * j.VPO4) / 3
                    r5 = (i.assignment * j.VPO5) / 3
                    r6 = (i.assignment * j.VPO6) / 3
                    r7 = (i.assignment * j.VPO7) / 3
                    r8 = (i.assignment * j.VPO8) / 3
                    r9 = (i.assignment * j.VPO9) / 3
                    r10 = (i.assignment * j.VPO10) / 3
                    r11 = (i.assignment * j.VPO11) / 3
                    r12 = (i.assignment * j.VPO12) / 3
                    
                    cd = Employee.objects.get(course=code)
                    name = cd.CourseName
                    check = Weights.objects.get(Code_patt_id=code)
                    if i.assignment >= check.target:
                        yn = 'YES'
                    else:
                        yn = 'NO'
                    b = Report(Course_name=name, assessment=i.assessment, CO_id=j.CO_id, Y_N=yn, VPO1=r1, VPO2=r2, VPO3=r3, VPO4=r4, VPO5=r5,
                               VPO6=r6, VPO7=r7, VPO8=r8, VPO9=r9, VPO10=r10, VPO11=r11, VPO12=r12)
                    b.save()

    return redirect('/showreport')


def showreport(request):
    obj = Report.objects.all()
    s1 = request.session['course']
    print(obj)
    return render(request, "report.html", {'obj': obj, 's1': s1})
    
    
def gencis(request):
    global target
    obj = []
    res = []
    k = 0
    s1 = request.session['course']
    code = s1[:4]
    pattern = s1[5:-1]
    course = Employee.objects.get(course=s1)

    print(s1)
    dm = DeliveryMethods.objects.get(Code_patt=s1)
    wt = Weights.objects.filter(Code_patt=s1)
    for z in wt:
        target = z.target
    a = b = c = d = e = f = g = h = 0
    if dm.chalktalk:
        a = 1
    if dm.ict:
        b = 1
    if dm.gd:
        c = 1
    if dm.ifv:
        d = 1
    if dm.et:
        e = 1
    if dm.sur:
        f = 1
    if dm.mp:
        g = 1
    if dm.lab:
        h = 1
    data = {
        'a': a,
        'b': b,
        'c': c,
        'd': d,
        'e': e,
        'f': f,
        'g': g,
        'h': h,
    }
    cnt = 0
    co = CO.objects.filter(CO_id__icontains=s1)
    for x in co:
        cnt = cnt + 1

    pos = PO.objects.all()
    map = CoPoMapp.objects.all()
    ass = Assesment.objects.all()
    j = 1
    for i in map:
        if j > cnt:
            break
        cmp = s1 + '.' + str(j)
        if i.CO_id_id == cmp:
            obj.append(i)
            j = j + 1

    j = 1
    for i in ass:
        if j > cnt:
            break
        cmp = s1 + '.' + str(j)
        if i.CO_id_id == cmp:
            res.append(i)
            j = j + 1
    result1 = {
        'res': res,
        'target': target,
    }
    es = ExamScheme.objects.get(Code_patt=s1)
    ts = TeachingScheme.objects.get(Code_patt=s1)
    return render(request, "cis.html",
                  {'data': data, 'course': course, 'result1': result1, 'obj': obj, 'co': co, 'pos': pos, 'es': es, 'ts': ts, 'code': code, 'pattern': pattern})
